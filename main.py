from PySide6.QtGui import QCloseEvent
import sys, os, json, zipfile, tempfile, random, threading, shutil, time, re
from datetime import datetime
import openpyxl
import requests
import pandas as pd
from openpyxl import Workbook
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel, QLineEdit, QPushButton,
    QFileDialog, QHBoxLayout, QVBoxLayout, QMessageBox, QProgressBar, QDialog,
    QFormLayout
)
from PySide6.QtCore import Signal
from PySide6.QtGui import QIcon

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


from order_processor import process_order_folder, is_confirmed_excel
import subprocess

import gspread
import google.auth
import google.auth.transport.requests
import google.oauth2.service_account
from google.oauth2.service_account import Credentials

from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from pathlib import Path
from random import randint
import traceback

SHEET_ID_MASTER = "1-HB7z7TmWoBhXPCXjp32biuYKB4ITxQfwdhQ_dO52l4" #메인 시트
#SHEET_ID_MASTER = "18JG34ZOg1VyWeQQTz4vA3M9fh1GkjFBfD3xUfV9XBOM" #회사 내부용 시트

STOCK_SHEET_CSV = (
    f"https://docs.google.com/spreadsheets/d/{SHEET_ID_MASTER}/export"
    "?format=csv&gid=679677814" 
)


#Google Credential Files
def load_credentials():
    if getattr(sys, 'frozen', False):
        # PyInstaller 실행 중
        base_dir = sys._MEIPASS
    else:
        base_dir = os.path.dirname(__file__)
    cred_path = os.path.join(base_dir, "google_credentials.json")
    with open(cred_path, "r", encoding="utf-8") as f:
        return json.load(f)

GOOGLE_CREDENTIALS_DICT = load_credentials()

_GSP_CLIENT = None 

def get_gspread_client():
    """gspread.Client를 1회만 초기화해 재사용"""
    global _GSP_CLIENT
    if _GSP_CLIENT is None:               # 아직 없으면 → 생성
        GOOGLE_CREDENTIALS_DICT["private_key"] = GOOGLE_CREDENTIALS_DICT["private_key"].replace("\\n", "\n")
        creds = Credentials.from_service_account_info(
            GOOGLE_CREDENTIALS_DICT,
            scopes=["https://www.googleapis.com/auth/spreadsheets"]
        )
        _GSP_CLIENT = gspread.authorize(creds)
    return _GSP_CLIENT

_DRIVE_SERVICE = None

def get_drive_service():
    global _DRIVE_SERVICE
    if _DRIVE_SERVICE is None:
        GOOGLE_CREDENTIALS_DICT["private_key"] = GOOGLE_CREDENTIALS_DICT["private_key"].replace("\\n", "\n")
        creds = Credentials.from_service_account_info(
            GOOGLE_CREDENTIALS_DICT,
            scopes=["https://www.googleapis.com/auth/drive"]
        )
        _DRIVE_SERVICE = build("drive", "v3", credentials=creds)
    return _DRIVE_SERVICE
# ─── 상수 ─────────────────────────────────────────────────────
CONFIG_FILE   = "config.json"

if getattr(sys, 'frozen', False):  # PyInstaller 실행 여부
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(__file__)

PRODUCT_XLSX = os.path.join(BASE_DIR, "상품정보.xlsx")

PRODUCT_HEADERS = [
    "상품바코드", "상품바코드명", "상품코드",
    "상품옵션1(중문)", "상품옵션2(중문)", "상품옵션3(중문)",
    "상품단가(위안)", "이미지URL", "상품URL",
    "통관품목명(영문)", "통관품목명(한글)",
    "소재(바코드표시)", "주의사항(바코드표시)",
    "포장1개당구매수량", "합포장여부", "메모"
]


ICON_PATH = os.path.join(os.path.dirname(__file__), "images", "cashbot.ico")

def find_column(df: pd.DataFrame, keywords: list) -> str | None:
    # 공백 제거 후 소문자 비교
    df_columns_cleaned = {col: col.strip().replace(" ", "").lower() for col in df.columns}
    for keyword in keywords:
        keyword_clean = keyword.strip().replace(" ", "").lower()
        for col, clean_col in df_columns_cleaned.items():
            if keyword_clean in clean_col:
                return col
    return None

def create_drive_folder(folder_name, parent_id=None):
    # 고정된 공유 폴더 ID 반환
    return "0AIUiN0FF2S3SUk9PVA"

def upload_folder_to_drive(folder_path, drive_folder_id):
    service = get_drive_service()   # 재사용!

    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        if not os.path.isfile(file_path):
            continue

        file_metadata = {
            "name": filename,
            "parents": [drive_folder_id],
        }

        media = MediaFileUpload(file_path, resumable=True)
        uploaded = service.files().create(
            body=file_metadata,
            media_body=media,
            fields="id",
            supportsAllDrives=True  # ✅ 이거 추가
        ).execute()
        print(f"✔ 업로드 완료: {filename} → Drive File ID: {uploaded['id']}")

def safe_strip(value):
    """None 또는 NaN을 안전하게 처리하여 문자열로 반환"""
    if pd.isna(value) or value is None:
        return ""
    return str(value).strip()

def load_purchase_price_map(list_path: str) -> dict[str, str]:
    """
    발주서리스트_*.xlsx → {바코드: 매입가}
    병합 헤더 완벽 대응 최종 버전
    """
    # 병합 헤더를 다중 헤더로 읽기 (엑셀 기준 20, 21번째 줄 → header=[19, 20])
    df_raw = pd.read_excel(list_path, dtype=str, header=[19, 20]).fillna("")

    # 열 이름이 MultiIndex로 들어옴 → 공백 제거 후 하나의 문자열로 합치기
    df_raw.columns = [' '.join(str(s).strip() for s in col if str(s).strip()) for col in df_raw.columns]

    # 바코드 열 찾기
    col_bar_candidates = [c for c in df_raw.columns if "barcode" in c.replace(" ", "").lower() or "바코드" in c]
    if not col_bar_candidates:
        raise Exception(f"발주서리스트 파일에서 바코드 열을 찾을 수 없습니다.\n현재 열: {df_raw.columns.tolist()}")
    col_bar = col_bar_candidates[0]

    # 매입가 열 찾기 (왼쪽 매입가 우선)
    cost_cols = [c for c in df_raw.columns if "매입가" in c.replace(" ", "")]
    if not cost_cols:
        raise Exception(f"발주서리스트 파일에서 '매입가' 열을 찾을 수 없습니다.\n현재 열: {df_raw.columns.tolist()}")
    col_cost = cost_cols[0]

    price_map = {}

    rows = df_raw[col_bar].tolist()
    costs = df_raw[col_cost].tolist()

    i = 0
    while i < len(rows) - 1:
        name = str(rows[i]).strip()
        barcode = str(rows[i + 1]).strip()
        purchase = str(costs[i]).strip()

        if barcode.startswith("R"):
            price_map[barcode] = purchase
            i += 2
        else:
            i += 1

    return price_map


def load_stock_df(biz_num: str, save_excel: bool = True) -> pd.DataFrame:
    try:
        client = get_gspread_client()
        sheet = client.open_by_key(SHEET_ID_MASTER)

        # ─────────────────────────────
        # ✅ 재고 리스트 처리
        ws_stock = sheet.worksheet("재고 리스트")
        data_stock = ws_stock.get_all_values()
        header = data_stock[0]
        records = data_stock[1:]

        print(f"[DEBUG] 열 개수: {len(header)}")

        df_stock = pd.DataFrame(records, columns=header).fillna("")

        # 열 이름 유연하게 찾기
        def find_column(possible_names: list[str]) -> str | None:
            for key in possible_names:
                for col in df_stock.columns:
                    if key.strip().lower() in col.strip().lower():
                        return col
            return None

        sku_col  = find_column(["SKU", "상품코드"])
        name_col = find_column(["제품명", "상품명"])
        bc_col   = find_column(["바코드", "barcode"])
        qty_col  = find_column(["수량", "재고", "재고수량"])
        biz_col  = find_column(["사업자 번호", "사업자", "사업자등록번호"])

        if not all([sku_col, name_col, bc_col, qty_col, biz_col]):
            print("[재고 시트 오류] 필수 열 누락 - SKU, 제품명, 바코드, 수량, 사업자번호 중 하나가 없습니다.")
            return pd.DataFrame(columns=["SKU", "상품명", "바코드", "수량"])

        df_filtered = df_stock[df_stock[biz_col].astype(str).str.strip() == biz_num]

        if df_filtered.empty:
            print(f"[INFO] 재고 시트에 해당 사업자번호 {biz_num} 에 대한 데이터 없음")
            return pd.DataFrame(columns=["SKU", "상품명", "바코드", "수량"])

        df_result = df_filtered[[sku_col, name_col, bc_col, qty_col]]
        df_result.columns = ["SKU", "상품명", "바코드", "수량"]

        # ─────────────────────────────
        # ✅ 저장: 재고 + 입출고
        if save_excel:
            if getattr(sys, 'frozen', False):
                # PyInstaller 실행 중
                base_dir = Path(sys.executable).parent
            else:
                base_dir = Path(__file__).parent

            save_dir = base_dir
            save_dir.mkdir(parents=True, exist_ok=True)

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            rand_suffix = randint(1000, 9999)

            stock_path = save_dir / f"재고_{biz_num}_{ts}_{rand_suffix}.xlsx"
            df_result.to_excel(stock_path, index=False)
            print(f"[INFO] 재고 저장 완료: {stock_path}")

            # ─────────────────────────────
            # ✅ 입출고 리스트 처리
            try:
                ws_inout = sheet.worksheet("입출고 리스트")
                data_inout = ws_inout.get_all_values()
                df_inout = pd.DataFrame(data_inout[1:], columns=data_inout[0]).fillna("")

                biz_col_io = next((c for c in df_inout.columns if "사업자 번호" in c), None)

                if biz_col_io:
                    df_filtered_io = df_inout[df_inout[biz_col_io].astype(str).str.strip() == biz_num]

                    if not df_filtered_io.empty:
                        io_path = save_dir / f"입출고리스트_{biz_num}_{ts}_{rand_suffix}.xlsx"
                        df_filtered_io.to_excel(io_path, index=False)
                        print(f"[INFO] 입출고리스트 저장 완료: {io_path}")
                else:
                    print("[INFO] 입출고리스트에서 '사업자 번호' 열을 찾지 못했습니다.")
            except Exception as e_io:
                print(f"[WARN] 입출고리스트 시트 처리 중 오류: {e_io}")

        return df_result

    except Exception as e:
        print("[load_stock_df 예외 발생]", type(e), e)
        traceback.print_exc()
        return pd.DataFrame(columns=["SKU", "상품명", "바코드", "수량"])


# ─── 설정 다이얼로그 ─────────────────────────────────────────
class SettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("쿠팡 ID/PW 설정")
        self.setFixedSize(320, 200)
        self.setWindowIcon(QIcon(ICON_PATH))
        lay = QFormLayout(self)
        
        self.le_biz = QLineEdit(); lay.addRow("사업자번호:", self.le_biz)
        self.le_id = QLineEdit();  lay.addRow("쿠팡 아이디:", self.le_id)
        self.le_pw = QLineEdit();  self.le_pw.setEchoMode(QLineEdit.Password)
        lay.addRow("쿠팡 비밀번호:", self.le_pw)
        self.le_brand = QLineEdit(); lay.addRow("브랜드명:", self.le_brand)

        btn = QPushButton("저장"); btn.clicked.connect(self._save); lay.addWidget(btn)
        self._load()

    def _load(self):
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                d = json.load(f)
            self.le_biz.setText(d.get("business_number", "")) 
            self.le_id.setText(d.get("coupang_id", ""))
            self.le_pw.setText(d.get("coupang_pw", ""))
            self.le_brand.setText(d.get("brand_name", ""))

    def _save(self):
        if not self.le_id.text().strip() or not self.le_pw.text().strip():
            QMessageBox.warning(self, "경고", "쿠팡 ID/PW를 입력하세요."); return
        if not self.le_biz.text().strip():
            QMessageBox.warning(self, "경고", "사업자번호를 입력하세요."); return
        data = {
            "business_number": self.le_biz.text().strip(), 
            "coupang_id": self.le_id.text().strip(),
            "coupang_pw": self.le_pw.text().strip(),
            "brand_name": self.le_brand.text().strip(),
        }
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        self.accept()
        parent = self.parent()
        if parent and hasattr(parent, "_enable_run"):
            parent._load_config()  


class UpdateWindow(QWidget):
    progressChanged = Signal(int)  # ✅ 시그널 정의

    def __init__(self, update_url, parent=None):
        super().__init__(parent)
        self.setWindowTitle("업데이트 중...")
        self.setFixedSize(300, 100)
        self.progress = QProgressBar(self)
        self.progress.setRange(0, 100)
        
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("업데이트 중입니다. 잠시만 기다려주세요."))
        layout.addWidget(self.progress)

        self.progressChanged.connect(self.progress.setValue)  # ✅ 시그널 → UI 연결

        self.show()

        # 업데이트 쓰레드 실행
        threading.Thread(target=self.perform_update_auto, args=(update_url,), daemon=True).start()

    def perform_update_auto(self, update_url):
        try:
            # 현재 exe가 있는 디렉토리 기준으로 다운로드 및 압축 해제
            base_dir = os.path.dirname(sys.argv[0])
            zip_path = os.path.join(base_dir, "balzubot_update.zip")
            extract_dir = os.path.join(base_dir, "balzubot_new")

            with requests.get(update_url, stream=True) as r:
                r.raise_for_status()
                total = int(r.headers.get('content-length', 0))
                downloaded = 0
                with open(zip_path, "wb") as f:
                    for chunk in r.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
                            downloaded += len(chunk)
                            percent = int(downloaded / total * 100)
                            self.progressChanged.emit(min(percent, 99))

            # 압축 해제
            if os.path.exists(extract_dir):
                shutil.rmtree(extract_dir)
            shutil.unpack_archive(zip_path, extract_dir)

            # 실행 파일 찾기 및 실행
            exe_files = [f for f in os.listdir(extract_dir) if f.endswith(".exe")]
            if not exe_files:
                self._show_error("실행 파일을 찾을 수 없습니다.")
                return

            self.progressChanged.emit(100)
            time.sleep(0.5)

            exe_path = os.path.join(extract_dir, exe_files[0])
            subprocess.Popen([exe_path])
            time.sleep(1)
            os._exit(0)

        except Exception as e:
            self._show_error(f"업데이트 실패: {e}")
            os._exit(1)

    def _show_error(self, msg):
        # ✅ 메시지박스는 메인 스레드에서 실행되도록 시그널로 처리해도 좋지만, 여기선 최소화 위해 직접 사용
        QMessageBox.critical(self, "업데이트 오류", msg)

    def closeEvent(self, event):
        os._exit(1)


# ─── 메인 윈도우 ────────────────────────────────────────────
class OrderApp(QMainWindow):

    crawlFinished   = Signal(str)
    crawlError      = Signal(str)
    progressUpdated = Signal(int)

    def __init__(self):
        super().__init__()

        self.skip_inventory_check = False
        
        self.business_number = ""

        self.setWindowTitle("수강생 발주 프로그램")
        self.setFixedSize(680, 300)
        self.setWindowIcon(QIcon(ICON_PATH))

        # 설정값
        self.order_zip_path = None
        self.coupang_id = self.coupang_pw = ""
        self.brand_name = ""

        # 런타임
        self.orders_data = {}
        self.cached_shipment = {}
        self.driver = None

        self.processed_files = set()  # ✅ 이미 처리한 파일 캐시
        self.cached_stock_df = None   # ✅ 재고 데이터 캐시

        self._build_ui(); self._load_config()
        self.progressUpdated.connect(lambda v: self.progress.setValue(v))
        self.crawlFinished.connect(self._crawl_ok)
        self.crawlError.connect(self._crawl_err)


    # UI ----------------------------------------------------------------
    def _build_ui(self):
        cen = QWidget(); self.setCentralWidget(cen)
        lay = QVBoxLayout(cen)

        # ZIP
        row_zip = QHBoxLayout(); row_zip.addWidget(QLabel("폴더 경로:"))
        self.le_zip = QLineEdit(); self.le_zip.setReadOnly(True)
        btn_zip = QPushButton("폴더 선택")
        btn_zip.clicked.connect(self._pick_order_folder)
        row_zip.addWidget(self.le_zip); row_zip.addWidget(btn_zip)

        # 브랜드
        row_brand = QHBoxLayout(); row_brand.addWidget(QLabel("브랜드명:"))
        self.le_brand = QLineEdit(); row_brand.addWidget(self.le_brand)

        # 설정
        row_set = QHBoxLayout(); row_set.addStretch()
        btn_set = QPushButton("쿠팡 ID/PW 설정"); btn_set.clicked.connect(self._open_settings)
        row_set.addWidget(btn_set)

        # 실행
        row_run = QHBoxLayout()
        self.btn_run = QPushButton("일괄 처리"); self.btn_run.clicked.connect(self._run_pipeline)
        self.btn_run.setEnabled(False); row_run.addWidget(self.btn_run)
        self.btn_batch = self.btn_run

        row_download = QHBoxLayout()
        btn_download = QPushButton("재고/입출고 다운로드")
        btn_download.clicked.connect(self._download_stock_io)
        row_download.addWidget(btn_download)
        lay.addLayout(row_download)

        # progress
        self.progress = QProgressBar(); self.progress.setRange(0, 100); self.progress.setVisible(False)

        for r in (row_zip, row_brand, row_set, row_run): lay.addLayout(r)
        lay.addWidget(self.progress)

        for w in (self.le_zip, self.le_brand): w.textChanged.connect(self._enable_run)


    def _download_stock_io(self):
        if not self.business_number:
            QMessageBox.warning(self, "사업자번호 없음", "먼저 설정에서 사업자번호를 입력하세요.")
            return
        try:
            result_df = load_stock_df(self.business_number, save_excel=True) 
            if result_df.empty:
                QMessageBox.information(self, "완료", "해당 사업자의 재고 데이터가 없습니다.")
            else:
                QMessageBox.information(self, "완료", "재고 및 입출고 리스트 다운로드가 완료되었습니다.")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"다운로드 중 오류 발생: {e}")
            
    def _enable_run(self):
        self.btn_run.setEnabled(bool(self.le_zip.text() and self.le_brand.text() and self.business_number))

    # 설정 로드 ----------------------------------------------------------
    def _load_config(self):
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                d = json.load(f)
            self.coupang_id = d.get("coupang_id", "")
            self.coupang_pw = d.get("coupang_pw", "")
            self.brand_name = d.get("brand_name", "")
            self.business_number = d.get("business_number", "")
            self.le_brand.setText(self.brand_name)
        self._enable_run()

    # UI slots ----------------------------------------------------------
    def _pick_order_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "발주서 폴더 선택")
        if folder:
            self.order_zip_path = folder  # 변수명 그대로 사용해도 무방
            self.le_zip.setText(folder)

    def _open_settings(self):
        if SettingsDialog(self).exec() == QDialog.Accepted:
            self._load_config()

    # ──────────────────────────────────────────────────────────
    # 파이프라인 시작
    # ──────────────────────────────────────────────────────────
    def _run_pipeline(self):
        try:
            # 1) 상품정보 템플릿 확인 ────────────────
            if not os.path.exists(PRODUCT_XLSX):
                wb = Workbook(); ws = wb.active
                ws.title = "상품정보"; ws.append(PRODUCT_HEADERS)
                wb.save(PRODUCT_XLSX)
                QMessageBox.information(
                    self, "상품정보 템플릿 생성",
                    "상품정보.xlsx 파일이 없어 템플릿을 생성했습니다.\n"
                    "정보를 채워 넣은 뒤 다시 실행해 주세요."
                )
                return

            # 2) 미확정 발주서만 복사해 임시 폴더 생성 ─
            if not self._zero_phase():
                return                      # 실패 시 바로 종료

            # 3) 발주확정·쉽먼트 양식 생성 ─────────────
            result = process_order_folder(self._temp_dir)

            # 4) 결과 알림 ────────────────────────────
            if result["failures"]:
                QMessageBox.warning(
                    self, "실패",
                    "처리 실패 파일:\n\n" + "\n".join(result["failures"])
                )
            else:
                QMessageBox.information(
                    self, "완료",
                    "파일 생성 완료!"
                )

            # 5) 다음 단계로 바로 진행  ❗❗
            self._first_phase()             # ← 두 번째 _zero_phase() 삭제

        except Exception as e:
            QMessageBox.critical(self, "오류", f"처리 중 오류:\n{e}")

        finally:
            # 임시 폴더 깔끔하게 삭제(선택) -------------
            try:
                if hasattr(self, "_temp_dir") and os.path.isdir(self._temp_dir):
                    shutil.rmtree(self._temp_dir)
            except Exception:
                pass

    # 0) ZIP 전처리 ------------------------------------------------------
    def _zero_phase(self):
        try:
            self.processed_files.clear()  # ✅ 새 발주 시작 시 캐시 초기화
            excel_files = []
            confirmed_skipped = 0

            for root, _, files in os.walk(self.order_zip_path):
                for fname in files:
                    if fname.lower().endswith((".xls", ".xlsx")):
                        full_path = os.path.join(root, fname)
                        if is_confirmed_excel(full_path):
                            confirmed_skipped += 1
                            continue
                        excel_files.append(full_path)

            if not excel_files:
                msg = (
                    f"모든 엑셀 파일이 발주 확정본으로 제외되었습니다. ({confirmed_skipped}건)"
                    if confirmed_skipped > 0 else "미확정 발주서가 없습니다."
                )
                QMessageBox.information(self, "안내", msg)
                return False

            # 선택된 엑셀 파일들을 임시 폴더에 복사 (기존 흐름 유지)
            self._temp_dir = tempfile.mkdtemp(prefix="order_folder_")
            for src in excel_files:
                shutil.copy2(src, self._temp_dir)

            return True

        except Exception as e:
            print("Zero Phase 오류:", e)
            return False

    # 1) 발주서 파싱 + 바코드 검증 + Selenium --------------------------------
    def _first_phase(self):
        try:
            print("[first_phase] 시작")

            excel_files = []
            for fname in os.listdir(self._temp_dir):
                if fname.lower().endswith((".xls", ".xlsx")):
                    excel_files.append(os.path.join(self._temp_dir, fname))

            if not excel_files:
                raise Exception("엑셀 파일이 없습니다.")

            self.list_path = next(
                (p for p in excel_files if "발주서리스트" in os.path.basename(p)),
                None
            )
            self.price_map = {}

            for p in excel_files:
                if "발주서리스트" in os.path.basename(p):
                    partial_map = load_purchase_price_map(p)
                    self.price_map.update(partial_map)

            for idx, xlsx in enumerate(excel_files):
                print(f"[first_phase] 처리 중: {os.path.basename(xlsx)}")

                # ✅ 캐시된 파일은 건너뜀
                if xlsx in self.processed_files:
                    print(f"[first_phase] 이미 처리된 파일, 건너뜀: {os.path.basename(xlsx)}")
                    continue

                try:
                    df_raw = pd.read_excel(xlsx, header=None, dtype=str)

                    po_row_series = df_raw[df_raw.iloc[:, 0].astype(str).str.contains("발주번호", na=False)].index
                    if po_row_series.empty:
                        raise ValueError(f"{os.path.basename(xlsx)} 파일에 '발주번호'가 없습니다.")
                    po_row = po_row_series[0]
                    po_no = str(df_raw.iloc[po_row, 2]).strip()

                    eta_row_series = df_raw[df_raw.iloc[:, 0].astype(str).str.contains("입고예정일시", na=False)].index
                    if eta_row_series.empty:
                        raise ValueError(f"{os.path.basename(xlsx)} 파일에 '입고예정일시'가 없습니다.")
                    eta_row = eta_row_series[0] + 1

                    eta_raw = df_raw.iloc[eta_row, 5]
                    eta = pd.to_datetime(eta_raw, errors="coerce")
                    if pd.isna(eta):
                        raise ValueError(f"입고예정일시 변환 실패: {eta_raw}")
                    eta = eta.to_pydatetime()

                    center = str(df_raw.iloc[eta_row, 2]).strip()

                    df_items = pd.read_excel(xlsx, header=19, dtype=str).fillna("")
                    df_items = df_items.loc[:, ~df_items.columns.str.startswith("Unnamed")]
                    df_items.columns = df_items.columns.str.strip()

                    col_barcode = next((c for c in df_items.columns if "BARCODE" in c.upper() or "바코드" in c), None)
                    if not col_barcode:
                        continue

                    rows = df_items[col_barcode].tolist()
                    valid_pairs = []
                    i = 0
                    while i < len(rows) - 1:
                        name = str(rows[i]).strip()
                        barcode = str(rows[i + 1]).strip()
                        if barcode.startswith("R"):
                            valid_pairs.append((name, barcode))
                            i += 2
                        else:
                            i += 1

                    for product_name, barcode in valid_pairs:
                        if not barcode:
                            continue

                        if po_no not in self.orders_data:
                            self.orders_data[po_no] = {
                                "barcode": barcode,
                                "product_code": "",
                                "product_name": product_name,
                                "center": center,
                                "eta": eta,
                                "shipment": None,
                                "invoice": str(random.randint(10**9, 10**10 - 1))
                            }

                    # ✅ 여기 추가
                    self.processed_files.add(xlsx)

                    pct = int((idx + 1) / len(excel_files) * 30)
                    self.progressUpdated.emit(pct)

                except Exception as e:
                    raise Exception(f"{os.path.basename(xlsx)} 파일 처리 중 오류: {e}")

            print("[first_phase] 상품정보 바코드 확인 시작")

            prod_df = pd.read_excel(PRODUCT_XLSX, dtype=str).fillna("")
            if "상품바코드" not in prod_df.columns:
                raise Exception("상품정보.xlsx에 '상품바코드' 열이 없습니다.")

            known_barcodes = set(prod_df["상품바코드"].astype(str).str.strip().str.lower())
            new_barcodes = []

            for xlsx in excel_files:
                df_items = pd.read_excel(xlsx, header=19, dtype=str).fillna("")
                df_items = df_items.loc[:, ~df_items.columns.str.startswith("Unnamed")]
                df_items.columns = df_items.columns.str.strip()

                col_barcode = next((c for c in df_items.columns if "BARCODE" in c.upper() or "바코드" in c), None)
                if not col_barcode:
                    continue

                rows = df_items[col_barcode].tolist()
                valid_pairs = []
                i = 0
                while i < len(rows) - 1:
                    name = str(rows[i]).strip()
                    barcode = str(rows[i + 1]).strip()
                    if barcode.startswith("R"):
                        valid_pairs.append((name, barcode))
                        i += 2
                    else:
                        i += 1

                for product_name, barcode in valid_pairs:
                    if not barcode:
                        continue
                    bc_lower = barcode.lower()
                    if bc_lower not in known_barcodes:
                        new_barcodes.append((barcode, product_name))

            added = set()
            rows_to_append = []
            for barcode, name in new_barcodes:
                bc_lower = barcode.lower()
                if bc_lower not in added:
                    row = [barcode, name, ""] + [""] * (len(PRODUCT_HEADERS) - 3)
                    rows_to_append.append(row)
                    added.add(bc_lower)

            if rows_to_append:
                wb = openpyxl.load_workbook(PRODUCT_XLSX)
                ws = wb.active
                for row in rows_to_append:
                    ws.append(row)
                wb.save(PRODUCT_XLSX)

                self.skip_inventory_check = True

                QMessageBox.information(
                    self, "상품정보 자동 추가",
                    f"{len(rows_to_append)}개 바코드를 상품정보.xlsx에 자동으로 추가했습니다.\n내용 확인 후 다시 실행해주세요."
                )
                return

            print("[first_phase] 재고 확인 시작")

            if not self.skip_inventory_check:
                try:
                    self.cached_stock_df = load_stock_df(self.business_number)  # ✅ 캐시에 저장
                    if self.cached_stock_df.empty:
                        QMessageBox.warning(self, "재고 시트 비어 있음", "현재 재고 시트에 데이터가 없습니다.\n계속 진행은 가능하지만 재고 확인은 생략됩니다.")
                except Exception as e:
                    QMessageBox.warning(self, "재고 확인 경고", f"재고 정보를 불러오는 중 오류 발생: {e}\n재고 확인을 생략하고 계속 진행합니다.")
            else:
                print("[first_phase] 재고 확인 스킵됨")

            print("[first_phase] Selenium 드라이버 시작")

            options = ChromeOptions()
            options.add_argument("--start-maximized")
            try:
                self.driver = webdriver.Chrome(options=options)
                print("[first_phase] 드라이버 실행 완료")
            except Exception as e:
                raise Exception(f"ChromeDriver 실행 실패: {e}")

            self.driver.implicitly_wait(5)
            oauth_url = (
                "https://xauth.coupang.com/auth/realms/seller/"
                "protocol/openid-connect/auth?response_type=code&client_id=supplier-hub"
                "&scope=openid&state=abc&redirect_uri=https://supplier.coupang.com/login/oauth2/code/keycloak"
            )
            try:
                self.driver.get(oauth_url)
                print("[first_phase] 로그인 페이지 로드 완료")
            except Exception as e:
                raise Exception(f"로그인 페이지 접속 실패: {e}")

            if self.coupang_id and self.coupang_pw:
                try:
                    WebDriverWait(self.driver, 15).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "input[name='username']"))
                    ).send_keys(self.coupang_id)
                    self.driver.find_element(By.CSS_SELECTOR, "input[name='password']").send_keys(self.coupang_pw)
                    self.driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
                except Exception as e:
                    raise Exception(f"쿠팡 로그인 실패: {e}")

            self.btn_batch.setText("로그인 완료")
            self.btn_batch.clicked.disconnect()
            self.btn_batch.clicked.connect(self.second_phase)
            self.btn_batch.setEnabled(True)

            print("[first_phase] 완료")

        except Exception as e:
            print("[예외 - first_phase]", repr(e))
            self.crawlError.emit(f"[first_phase 오류] {repr(e)}")

    # ──────────────────────────────────────────────────────────
    # 2) Selenium 로그인 완료 후 크롤링
    # ──────────────────────────────────────────────────────────
    def second_phase(self):
        self.btn_batch.setEnabled(False)
        self.progress.setVisible(True)
        threading.Thread(target=self.crawl_and_generate).start()

    def crawl_and_generate(self):
        try:
            driver = self.driver
            self.progressUpdated.emit(30)

            driver.get("https://supplier.coupang.com/dashboard/KR")

            # 2-1) Logistics → Shipments 메뉴 진입
            try:
                btn_logistics = WebDriverWait(driver, 15).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='/logistics']"))
                ); btn_logistics.click()

                btn_shipments = WebDriverWait(driver, 15).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='/ibs/asn/active']"))
                ); btn_shipments.click()
            except Exception:
                raise Exception("메뉴 클릭 실패 (Logistics → Shipments)")

            # 2-2) 발주번호 입력창 확인
            try:
                search_input = WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "input#purchaseOrderSeq"))
                )
            except:
                raise Exception("발주번호 입력창을 찾지 못했습니다.")

            # 2-3) 다운로드 폴더
            download_dir = os.path.join(os.path.expanduser("~"), "Downloads")
            target_dir   = os.path.join(os.getcwd(), "shipment"); os.makedirs(target_dir, exist_ok=True)

            # 2-4) 주문별 라벨/매니페스트 다운로드
            total = len(self.orders_data)
            for idx, (po_no, info) in enumerate(self.orders_data.items()):
                search_input.clear(); search_input.send_keys(po_no)
                driver.find_element(By.CSS_SELECTOR, "button#shipment-search-btn").click()

                try:
                    first_td = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR,
                            "table#parcel-tab tbody tr:first-child td:first-child"))
                    ); shipment_no = first_td.text.strip()
                except:
                    shipment_no = ""

                center, eta = info["center"], info["eta"]
                key = f"{center}|{eta.strftime('%Y-%m-%d') if eta else ''}"
                self.cached_shipment[key] = shipment_no
                self.orders_data[po_no]["shipment"] = shipment_no

                if shipment_no:
                    try:
                        driver.execute_script(
                            f"window.open('https://supplier.coupang.com/ibs/shipment/parcel/"
                            f"pdf-label/generate?parcelShipmentSeq={shipment_no}', '_blank');"
                        ); time.sleep(1.5)
                        driver.execute_script(
                            f"window.open('https://supplier.coupang.com/ibs/shipment/parcel/"
                            f"pdf-manifest/generate?parcelShipmentSeq={shipment_no}', '_blank');"
                        ); time.sleep(1.5)
                    except Exception as e:
                        print(f"[경고] {shipment_no} 다운로드 중 오류: {e}")

                percent = 30 + int((idx + 1) / total * 40)
                self.progressUpdated.emit(percent)

            # 2-5) 다운로드 완료 대기 후 정리
            time.sleep(5)
            dup_pat = re.compile(r"\s\(\d+\)(\.[^.]+)$")

            for fname in os.listdir(download_dir):
                low = fname.lower()
                if not (low.startswith("shipment_label_document") or
                        low.startswith("shipment_manifest_document")):
                    continue
                src = os.path.join(download_dir, fname)

                if dup_pat.search(fname):   # 중복본 삭제
                    try: os.remove(src)
                    except FileNotFoundError: pass
                    continue

                shutil.move(src, os.path.join(target_dir, fname))

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            drive_folder_name = f"shipment_{ts}"

            try:
                drive_folder_id = create_drive_folder("shipment")  # 아무 이름 넣어도 됨
                upload_folder_to_drive(target_dir, drive_folder_id)
                print(f"📁 Google Drive 업로드 완료: 공유폴더")
            except Exception as e:
                raise RuntimeError(f"Google Drive 업로드 실패: {e}") from e

            driver.quit(); self.driver = None
            self.progressUpdated.emit(100)
            self.crawlFinished.emit("전송 완료!")

        except Exception as e:
            print("crawl_and_generate 예외 발생:", e)
            self.crawlError.emit(str(e))

    # ──────────────────────────────────────────────────────────
    # 3) 3PL 신청서 & 주문서 생성
    # ──────────────────────────────────────────────────────────
    def generate_orders(self):

        def append_to_google_sheet(sheet_id: str, sheet_name: str, rows: list[list[str]]):
            client = get_gspread_client()
            ws = client.open_by_key(sheet_id).worksheet(sheet_name)

            content_rows = rows[1:]  # 헤더 제외
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for row in content_rows:
                row.append(now_str)
            
            ws.append_rows(content_rows, value_input_option="USER_ENTERED")

        try:
            # ─────────────────────────────────────────────
            # 0) 재고·확정 양식 로드
            # ─────────────────────────────────────────────
            # ✅ 캐시 우선 사용
            if self.cached_stock_df is not None:
                inv_df = self.cached_stock_df
            else:
                inv_df = load_stock_df(self.business_number, save_excel=False)

            inventory = {str(r["바코드"]).strip(): int(float(r["수량"] or 0))
                        for _, r in inv_df.iterrows()}
            used_stock = {}

            confirm_path = "발주 확정 양식.xlsx"
            df_confirm = pd.read_excel(confirm_path, dtype=str).fillna("")

            df_confirm["확정수량"] = pd.to_numeric(
                df_confirm["확정수량"], errors="coerce"
            ).fillna(0).astype(int)
            df_confirm["Shipment"] = df_confirm["발주번호"].map(
                lambda x: self.orders_data.get(str(x).strip(), {}).get("shipment", "")
            )
            df_confirm = df_confirm[df_confirm["확정수량"] > 0]

            price_map = getattr(self, "price_map", {})

            group_cols = ["Shipment", "상품바코드", "상품이름", "물류센터", "입고예정일"]
            df_group = (df_confirm[group_cols + ["확정수량"]]
                        .groupby(group_cols, as_index=False)["확정수량"].sum())

            brand = self.le_brand.text().strip()
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            biz_num = self.business_number.strip()

            hed_3pl = ["브랜드명", "쉽먼트번호", "발주번호", "SKU번호",
                    "SKU(제품명)", "바코드", "수량",
                    "입고예정일", "센터명", "사업자번호"]
            rows_3pl_file = [hed_3pl]
            rows_3pl_sheet = [hed_3pl + ["매입가"]]

            hed_ord = [
                "바코드명", "바코드", "상품코드", "쿠팡납품센터명",
                "쿠팡쉽먼트번호", "쿠팡입고예정일자", "입고마감준수여부",
                "발주수량", "중국재고사용여부", "제조일자", "유통기한"
            ]
            rows_order = [hed_ord]

            wb_3pl, ws_3pl = Workbook(), None
            ws_3pl = wb_3pl.active
            ws_3pl.title = "3PL신청서"
            ws_3pl.append(hed_3pl)

            wb_ord, ws_ord = Workbook(), None
            ws_ord = wb_ord.active
            ws_ord.title = "주문서"
            ws_ord.append(hed_ord)


            # 제조일자/유통기한/유통기한관리 매핑 준비 (발주서리스트에서 추출)
            mfg_map = {}  # {barcode: 제조일자}
            exp_map = {}  # {barcode: 유통기한}
            exp_flag_map = {}  # {barcode: 유통기한관리(Y/N)}
            # order_folder 내 모든 발주서리스트 파일을 순회
            order_folder = self.order_zip_path if hasattr(self, "order_zip_path") else None
            list_paths = []
            if order_folder and os.path.isdir(order_folder):
                for fname in os.listdir(order_folder):
                    if "발주서리스트" in fname and fname.lower().endswith((".xls", ".xlsx")):
                        list_paths.append(os.path.join(order_folder, fname))

            if list_paths:
                for list_path in list_paths:
                    try:
                        wb = openpyxl.load_workbook(list_path)
                        ws = wb.active
                        rows = list(ws.iter_rows(values_only=True))
                        # 21번째(엑셀 22행)부터 2줄씩
                        for i in range(21, len(rows), 2):
                            row1 = rows[i]
                            row2 = rows[i+1] if i+1 < len(rows) else None
                            if not row2:
                                continue
                            # 인덱스: row1[16]=관리, row1[17]=제조일자, row2[2]=바코드
                            barcode = row2[2] if len(row2) > 2 else None
                            mfg_date = row1[17] if len(row1) > 17 else None
                            exp_date = row1[17] if len(row1) > 17 else None
                            exp_flag = row1[16] if len(row1) > 16 else None
                            if barcode and str(barcode).startswith("R"):
                                if mfg_date:
                                    mfg_map[barcode] = mfg_date
                                if exp_date:
                                    exp_map[barcode] = exp_date
                                if exp_flag:
                                    exp_flag_map[barcode] = exp_flag
                    except Exception as e:
                        print(f"[WARN] 제조일자/유통기한/유통기한관리 추출 실패: {e} ({list_path})")
            else:
                print("[WARN] 발주서리스트 원본 파일을 찾을 수 없습니다.")


            for _, r in df_group.iterrows():
                bc = safe_strip(r["상품바코드"])
                pname = r["상품이름"]
                center = r["물류센터"]
                ship_no = r["Shipment"]
                eta_raw = r["입고예정일"]
                qty = int(r["확정수량"])
                eta_str = (pd.to_datetime(eta_raw, errors="coerce").strftime("%Y-%m-%d") if eta_raw else "")

                mask = (df_confirm["Shipment"] == ship_no) & (df_confirm["상품바코드"] == bc)
                po_no = product_code = ""
                if mask.any():
                    po_no = str(df_confirm.loc[mask, "발주번호"].iloc[0]).strip()
                    product_code = str(df_confirm.loc[mask, "상품번호"].iloc[0]).strip()

                row_base = [brand, ship_no, po_no, product_code,
                            pname, bc, qty, eta_str, center, biz_num]

                rows_3pl_file.append(row_base)
                ws_3pl.append(row_base)

                purchase = price_map.get(bc, "")
                rows_3pl_sheet.append(row_base + [purchase])

                already = used_stock.get(bc, 0)
                avail = inventory.get(bc, 0) - already
                need = max(qty - max(avail, 0), 0)

                # 제조일자/유통기한/유통기한관리 매핑 (바코드별 인덱스 매칭)
                def get_split_val(map_obj, bc):
                    return map_obj.get(bc, "")

                exp_flag_val = exp_flag_map.get(bc, "Y")
                has_y = str(exp_flag_val).strip().upper() == "Y"
                mfg_val = get_split_val(mfg_map, bc)
                exp_val = ""
                # 디버깅용 출력
                print(f"[DEBUG] bc={bc} exp_flag={exp_flag_val} mfg={mfg_val}")
                if has_y:
                    if mfg_val:
                        try:
                            mfg_dt = pd.to_datetime(mfg_val, errors="coerce")
                            if pd.notna(mfg_dt):
                                exp_val = "제조일로부터 1년"
                            else:
                                exp_val = ""
                        except Exception:
                            exp_val = ""
                # need > 0일 때만 주문서에 추가
                if need > 0:
                    row_ord = [pname, bc, product_code, center,
                            ship_no, eta_str, "Y", need, "N", mfg_val, exp_val]
                    rows_order.append(row_ord)
                    ws_ord.append(row_ord)

                used_stock[bc] = already + min(qty, max(avail, 0))

            append_to_google_sheet(
                sheet_id=SHEET_ID_MASTER,
                sheet_name="CALL 요청서",
                rows=rows_3pl_sheet
            )

            if len(rows_order) == 1:
                ws_ord.cell(row=2, column=1).value = "재고가 충분하여 주문할 항목이 없습니다."
                append_to_google_sheet(
                    sheet_id=SHEET_ID_MASTER,
                    sheet_name="CALL 주문서",
                    rows=[["재고가 충분하여 주문할 항목이 없습니다."]]
                )
            else:
                append_to_google_sheet(
                    sheet_id=SHEET_ID_MASTER,
                    sheet_name="CALL 주문서",
                    rows=rows_order
                )

            wb_3pl.save(f"3PL신청내역_{ts}.xlsx")
            wb_ord.save(f"주문서_{ts}.xlsx")

            QMessageBox.information(
                self, "완료",
                f"스프레드시트 전송 완료!\n"
                f"파일도 저장했습니다:\n"
                f"- 3PL신청내역_{ts}.xlsx\n"
                f"- 주문서_{ts}.xlsx"
            )

        except Exception as e:
            QMessageBox.critical(self, "오류", f"주문서 생성 중 오류:\n{e}")

    # ──────────────────────────────────────────────────────────
    # 크롤 완료/오류 콜백 및 버튼 리셋
    # ──────────────────────────────────────────────────────────
    def _crawl_ok(self, msg: str):
        self.progress.setVisible(False)

        try:
            self.generate_orders()               # ← 모든 실질 작업의 마지막 단계
            QMessageBox.information(             # 여기서 최종 알림
                self, "완료", "모든 작업이 끝났습니다!"
            )
        except Exception as e:
            QMessageBox.critical(self, "주문서 오류", str(e))

        self._reset_btn()

    def _crawl_err(self, msg: str):
        self.progress.setVisible(False)
        if not msg.strip():  # 빈 문자열이면
            msg = "에러 발생 (상세 메시지 없음)"
        QMessageBox.critical(self, "크롤 오류", msg)
        if self.driver:
            self.driver.quit()
            self.driver = None
        self._reset_btn()

    def _reset_btn(self):
        self.btn_run.setText("일괄 처리")
        self.btn_run.clicked.disconnect(); self.btn_run.clicked.connect(self._run_pipeline)
        self.btn_run.setEnabled(True)


# ─── main ───────────────────────────────────────────────────
if __name__ == "__main__":
    app = QApplication(sys.argv)

    try:
        VERSION_URL = "http://114.207.245.49/version"
        LOCAL_VERSION = "1.1.0"
        r = requests.get(VERSION_URL, timeout=5)
        if r.status_code == 200:
            data = r.json()
            latest_ver = data.get("balzubotversion")
            update_url = data.get("balzubot_update_url")
            if latest_ver and update_url and latest_ver != LOCAL_VERSION:
                update_win = UpdateWindow(update_url)
                sys.exit(app.exec())

    except Exception as e:
        QMessageBox.critical(None, "업데이트 오류", str(e))
        sys.exit(1)

    # 최신이면 본 프로그램 실행
    win = OrderApp()
    win.show()
    sys.exit(app.exec())