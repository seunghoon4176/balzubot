# file: order_processor.py

import os
import re
import zipfile
import shutil
import random
import sys
import pandas as pd
from openpyxl import load_workbook, Workbook
from decimal import Decimal, ROUND_HALF_UP
from typing import Optional


def round_to_hundred(x: float) -> int:
    return int(Decimal(x).quantize(Decimal('1E2'), rounding=ROUND_HALF_UP))


def get_output_dir() -> str:
    """
    PyInstaller로 묶였을 때는 sys.executable의 폴더를,
    일반 .py 실행 시에는 현재 작업 디렉터리(cwd)를 반환합니다.
    """
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    else:
        return os.getcwd()


def _find_cell_by_label(ws, label: str, max_row: int = 40, max_col: int = 15) -> Optional[str]:
    tgt = label.replace(" ", "")
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and tgt in v.replace(" ", ""):
                return ws.cell(row=r, column=c + 1).value
    return None


def restore_korean(name: str) -> str:
    try:
        return name.encode("cp437").decode("cp949")
    except UnicodeDecodeError:
        return name


def unzip_orders(zip_path: str) -> str:
    """
    get_output_dir() 하위의 'orders_unzip' 폴더에 ZIP 내부의 엑셀 파일들을 풀어둡니다.
    """
    out_dir = get_output_dir()
    dst = os.path.join(out_dir, "orders_unzip")
    if os.path.isdir(dst):
        shutil.rmtree(dst)
    os.makedirs(dst, exist_ok=True)

    with zipfile.ZipFile(zip_path) as zf:
        seen = set()
        for info in zf.infolist():
            if info.is_dir():
                continue
            if not info.filename.lower().endswith((".xlsx", ".xls", ".xlsm", ".xlsb")):
                continue

            fixed = restore_korean(os.path.basename(info.filename))
            base, ext = os.path.splitext(fixed)
            i = 1
            while fixed in seen:
                fixed = f"{base}_{i}{ext}"
                i += 1
            seen.add(fixed)

            out_path = os.path.join(dst, fixed)
            with zf.open(info) as src, open(out_path, "wb") as dst_f:
                shutil.copyfileobj(src, dst_f)

    return dst


def parse_orders(unzip_dir: str):
    def generate_invoice_number():
        return str(random.randint(100000000000, 999999999999))

    order_rec = []
    ship_rec = []
    inv_map = {}
    fails = []

    for root, _, files in os.walk(unzip_dir):
        for fname in sorted(files):
            if not fname.lower().endswith((".xls", ".xlsx", ".xlsm", ".xlsb")):
                continue

            wb_path = os.path.join(root, fname)
            ws = load_workbook(wb_path, data_only=True).active

            po = _find_cell_by_label(ws, "발주번호") or ws["C10"].value
            fc = ws["C13"].value
            edd_raw = _find_cell_by_label(ws, "입고예정일") or ws["F13"].value
            if not (po and fc and edd_raw):
                fails.append(fname)
                continue

            po, fc = str(po).strip(), str(fc).strip()
            edd = (
                edd_raw.strftime("%Y%m%d")
                if hasattr(edd_raw, "strftime")
                else re.sub(r"[^\d]", "", str(edd_raw))[:8]
            )
            if len(edd) != 8:
                fails.append(fname)
                continue

            key = (edd, fc)
            if key not in inv_map:
                inv_map[key] = generate_invoice_number()
            file_inv = inv_map[key]

            return_mgr = _find_cell_by_label(ws, "회송담당자") or ws["C14"].value
            return_tel = _find_cell_by_label(ws, "연락처") or ws["G14"].value
            return_addr = _find_cell_by_label(ws, "회송지") or ws["C15"].value

            header_row = next(
                (
                    i
                    for i in range(1, ws.max_row + 1)
                    if isinstance(ws.cell(row=i, column=3).value, str)
                    and "상품명" in ws.cell(row=i, column=3).value
                ),
                None,
            )
            if header_row is None:
                fails.append(fname)
                continue

            last_idx = None
            for r in range(header_row + 1, ws.max_row + 1):
                cell_val = ws.cell(row=r, column=3).value
                # 송장 바코드 푸는 부분 (R\d+ 패턴)
                if isinstance(cell_val, str) and re.match(r"^R\d+$", cell_val.strip()):
                    if last_idx is not None:
                        barcode = cell_val.strip()
                        order_rec[last_idx]["상품바코드"] = barcode
                        ship_rec[last_idx]["상품바코드(SKU Barcode)"] = barcode
                    continue

                sku = ws.cell(row=r, column=2).value
                if sku is None:
                    continue

                product_name = str(cell_val).strip() if cell_val else ""
                qty = ws.cell(row=r, column=7).value
                cost = ws.cell(row=r, column=10).value
                supply = ws.cell(row=r, column=11).value
                vat = ws.cell(row=r, column=12).value
                total = ws.cell(row=r, column=13).value

                order_rec.append(
                    {
                        "발주번호": po,
                        "물류센터": fc,
                        "입고유형": "쉽먼트",
                        "발주상태": "거래처확인요청",
                        "상품번호": sku,
                        "상품바코드": "",
                        "상품이름": product_name,
                        "발주수량": qty,
                        "확정수량": qty,
                        "유통(소비기한)": "",
                        "제조일자": "",
                        "생산년도": "",
                        "납품부족사유": "",
                        "회송담당자": return_mgr,
                        "회송담당자 연락처": return_tel,
                        "회송지주소": return_addr,
                        "매입가": cost,
                        "공급가": supply,
                        "부가세": vat,
                        "총발주매입금": total,
                        "입고예정일": edd,
                        "발주등록일시": "",
                    }
                )
                ship_rec.append(
                    {
                        "발주번호(PO ID)": po,
                        "물류센터(FC)": fc,
                        "입고유형(Transport Type)": "쉽먼트",
                        "입고예정일(EDD)": edd,
                        "상품번호(SKU ID)": sku,
                        "상품바코드(SKU Barcode)": "",
                        "상품이름(SKU Name)": product_name,
                        "확정수량(Confirmed Qty)": qty,
                        "송장번호(Invoice Number)": file_inv,
                        "납품수량(Shipped Qty)": qty,
                    }
                )
                last_idx = len(order_rec) - 1

    orders = pd.DataFrame(order_rec)
    ships = pd.DataFrame(ship_rec)

    # 컬럼 형식 보정
    orders["입고예정일"] = orders["입고예정일"].astype(str)
    ships["입고예정일(EDD)"] = ships["입고예정일(EDD)"].astype(str)
    ships["송장번호(Invoice Number)"] = ships["송장번호(Invoice Number)"].astype(str)
    ships[["Unnamed: 10", "주의사항"]] = ""

    return orders, ships, fails


def save_shipments(ship_df: pd.DataFrame):
    cols = [
        "발주번호(PO ID)",
        "물류센터(FC)",
        "입고유형(Transport Type)",
        "입고예정일(EDD)",
        "상품번호(SKU ID)",
        "상품바코드(SKU Barcode)",
        "상품이름(SKU Name)",
        "확정수량(Confirmed Qty)",
        "송장번호(Invoice Number)",
        "납품수량(Shipped Qty)",
        "Unnamed: 10",
        "주의사항",
    ]
    out_dir = get_output_dir()

    for edd, grp in ship_df.groupby("입고예정일(EDD)"):
        filename = f"쉽먼트 일괄 양식_{edd}.xlsx"
        path = os.path.join(out_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "상품목록"
        ws.append(cols)
        for row in grp.itertuples(index=False):
            ws.append(list(row))

        # “송장번호(Invoice Number)” 열을 텍스트 형식으로 지정
        idx = cols.index("송장번호(Invoice Number)") + 1
        for row_cells in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            for c in row_cells:
                c.number_format = "@"

        wb.create_sheet("송장번호입력")
        wb.create_sheet("입력방법")
        wb.save(path)


def process_order_zip(zip_path: str):
    unzip_dir = unzip_orders(zip_path)
    orders, ships, fails = parse_orders(unzip_dir)

    out_dir = get_output_dir()
    order_filename = "발주 확정 양식.xlsx"
    order_path = os.path.join(out_dir, order_filename)
    orders.to_excel(order_path, index=False)

    save_shipments(ships)

    shutil.rmtree(unzip_dir, ignore_errors=True)
    return {"failures": fails}
